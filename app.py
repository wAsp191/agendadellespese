import streamlit as st
import pandas as pd
import plotly.express as px
import os
from datetime import date
import io
from openpyxl.styles import Font, Alignment

# --- 1. SICUREZZA E CONFIGURAZIONE ---
PASSWORD_ACCESSO = "poggio2026" 

st.set_page_config(page_title="Bilancio Casa", layout="centered", page_icon="💰")

if "autenticato" not in st.session_state:
    st.session_state["autenticato"] = False

if not st.session_state["autenticato"]:
    st.title("🔐 Accesso Riservato")
    pwd = st.text_input("Inserisci la password di famiglia:", type="password")
    if st.button("Entra"):
        if pwd == PASSWORD_ACCESSO:
            st.session_state["autenticato"] = True
            st.rerun()
        else:
            st.error("Password errata!")
    st.stop()

# --- 2. LOGICA DATI E CATEGORIE ---
DATA_FILE = "spese_casa_v4.csv"
RECURRING_FILE = "modelli_ricorrenti_v4.csv"

color_map = {
    "Acqua": "#0000FF",         # Blu
    "Luce": "#FFFF00",          # Giallo
    "Gas": "#008000",           # Verde
    "Tari": "#000000",          # Nero
    "Tel/Internet": "#FFA500",  # Arancione
    "Alimenti": "#800080",      # Viola
    "Abbonamenti": "#808080",   # Grigio
    "Altro": "#A9A9A9"          # Grigio scuro
}
cat_lista_clean = list(color_map.keys())
cat_lista_filtro = ["Tutte"] + cat_lista_clean

def load_data(file, columns):
    if os.path.exists(file):
        try:
            df_loaded = pd.read_csv(file)
            if not df_loaded.empty:
                df_loaded['Data'] = pd.to_datetime(df_loaded['Data'], errors='coerce')
                df_loaded = df_loaded.dropna(subset=['Data'])
            return df_loaded
        except: return pd.DataFrame(columns=columns)
    return pd.DataFrame(columns=columns)

df = load_data(DATA_FILE, ["Data", "Categoria", "Descrizione", "Importo", "Periodo"])
df_rec = load_data(RECURRING_FILE, ["Categoria", "Descrizione", "Importo", "Cadenza", "Data"])

# --- 3. FUNZIONE EXCEL PRO ---
def to_excel_pro(df_to_download):
    output = io.BytesIO()
    df_excel = df_to_download.copy()
    df_excel['Data'] = df_excel['Data'].dt.strftime('%d/%m/%Y')
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_excel.to_excel(writer, index=False, sheet_name='Report Spese')
        workbook = writer.book
        worksheet = writer.sheets['Report Spese']
        bold_font = Font(bold=True, size=14)
        for row in worksheet.iter_rows(min_row=1, max_row=len(df_excel)+1):
            for cell in row:
                cell.alignment = Alignment(vertical='center', horizontal='left')
                if cell.row == 1: cell.font = bold_font
            worksheet.row_dimensions[row[0].row].height = 30
        for col in worksheet.columns:
            worksheet.column_dimensions[col[0].column_letter].width = 30
        last_row = len(df_excel) + 2
        worksheet.cell(row=last_row, column=3, value="TOTALE GENERALE:").font = bold_font
        worksheet.cell(row=last_row, column=4, value=df_to_download['Importo'].sum()).font = bold_font
        worksheet.row_dimensions[last_row].height = 40
    return output.getvalue()

# --- 4. SIDEBAR ---
with st.sidebar:
    st.header("💰 Opzioni")
    anni = sorted(df['Data'].dt.year.unique(), reverse=True) if not df.empty else [date.today().year]
    anno_sel = st.selectbox("Anno", anni)
    cat_sel = st.selectbox("Filtra Categoria", cat_lista_filtro)
    if not df.empty:
        df_f_excel = df[df['Data'].dt.year == anno_sel].copy()
        if cat_sel != "Tutte": df_f_excel = df_f_excel[df_f_excel["Categoria"] == cat_sel]
        if not df_f_excel.empty:
            st.download_button("📥 Report Excel", data=to_excel_pro(df_f_excel), 
                             file_name=f"Spese_{anno_sel}.xlsx", use_container_width=True)
    if st.button("Logout", use_container_width=True):
        st.session_state["autenticato"] = False
        st.rerun()

# Filtraggio dati
df_f = df[df['Data'].dt.year == anno_sel].copy() if not df.empty else pd.DataFrame()
if cat_sel != "Tutte" and not df_f.empty:
    df_f = df_f[df_f["Categoria"] == cat_sel]

# --- 5. UI PRINCIPALE ---
st.title("💰 Spese di Casa")

# AGGIUNGI SPESA
with st.expander("➕ **AGGIUNGI SPESA**", expanded=False):
    t1, t2 = st.tabs(["✍️ Manuale", "🔁 Modelli"])
    with t1:
        with st.form("fm", clear_on_submit=True):
            d = st.date_input("Data", date.today())
            c = st.selectbox("Categoria", cat_lista_clean)
            i = st.number_input("Importo (€)", min_value=0.0, step=0.01)
            p = st.text_input("Periodo (es. Gen-Feb)")
            ds = st.text_input("Nota")
            if st.form_submit_button("SALVA", use_container_width=True):
                nuova = pd.DataFrame([[pd.to_datetime(d), c, ds, i, p]], columns=df.columns)
                df = pd.concat([df, nuova], ignore_index=True); df.to_csv(DATA_FILE, index=False); st.rerun()
    with t2:
        if df_rec.empty: st.info("Configura modelli in fondo.")
        for idx, row in df_rec.iterrows():
            if st.button(f"📌 {row['Descrizione']} (€{row['Importo']})", key=f"r_{idx}", use_container_width=True):
                nuova_s = pd.DataFrame([[pd.to_datetime(date.today()), row['Categoria'], row['Descrizione'], row['Importo'], "Ricorrente"]], columns=df.columns)
                df = pd.concat([df, nuova_s], ignore_index=True); df.to_csv(DATA_FILE, index=False); st.rerun()

# DASHBOARD
if not df_f.empty:
    st.metric(f"Totale {cat_sel}", f"€ {df_f['Importo'].sum():,.2f}")
    
    # TORTA
    df_pie = df_f.groupby('Categoria')['Importo'].sum().reset_index()
    df_pie['Etichetta'] = df_pie.apply(lambda r: f"{r['Categoria']}: €{r['Importo']:,.2f}", axis=1)
    fig_pie = px.pie(df_pie, values='Importo', names='Etichetta', hole=0.5, title="Suddivisione", color='Categoria', color_discrete_map=color_map)
    st.plotly_chart(fig_pie, use_container_width=True)
    
    # BARRE IMPILATE (CALENDARIO COMPLETO)
    mesi_nomi = ["Gen", "Feb", "Mar", "Apr", "Mag", "Giu", "Lug", "Ago", "Set", "Ott", "Nov", "Dic"]
    df_m = df_f.copy()
    df_m['M_Num'] = df_m['Data'].dt.month
    res_m = df_m.groupby(['M_Num', 'Categoria'])['Importo'].sum().reset_index()
    
    fig_bar = px.bar(res_m, x='M_Num', y='Importo', color='Categoria', title=f"Andamento Annuale {anno_sel}", color_discrete_map=color_map)
    
    # Forza l'asse X a mostrare tutti i 12 mesi
    fig_bar.update_layout(
        xaxis=dict(tickmode='array', tickvals=list(range(1, 13)), ticktext=mesi_nomi, range=[0.5, 12.5]),
        barmode='stack', xaxis_title="Mesi dell'anno", yaxis_title="Spesa Totale (€)"
    )
    st.plotly_chart(fig_bar, use_container_width=True)

    # STORICO E MODIFICA/ELIMINAZIONE
    st.divider()
    st.subheader("📜 Gestione Spese")
    
    # Sezione Modifica/Elimina
    with st.expander("📝 MODIFICA O ELIMINA UNA VOCE"):
        # Seleziona la riga tramite un menu a tendina
        idx_to_edit = st.selectbox("Seleziona la spesa da gestire:", df_f.index, 
                                   format_func=lambda x: f"{df.loc[x,'Data'].strftime('%d/%m/%Y')} - {df.loc[x,'Descrizione']} (€{df.loc[x,'Importo']})")
        
        row_data = df.loc[idx_to_edit]
        
        # Form di modifica
        with st.form("edit_form"):
            col1, col2 = st.columns(2)
            new_date = col1.date_input("Modifica Data", row_data['Data'])
            new_cat = col1.selectbox("Modifica Categoria", cat_lista_clean, index=cat_lista_clean.index(row_data['Categoria']))
            new_imp = col2.number_input("Modifica Importo (€)", value=float(row_data['Importo']), step=0.01)
            new_per = col2.text_input("Modifica Periodo", value=row_data['Periodo'])
            new_des = st.text_input("Modifica Nota/Descrizione", value=row_data['Descrizione'])
            
            c_btn1, c_btn2 = st.columns(2)
            if c_btn1.form_submit_button("✅ AGGIORNA VOCE", use_container_width=True):
                df.at[idx_to_edit, 'Data'] = pd.to_datetime(new_date)
                df.at[idx_to_edit, 'Categoria'] = new_cat
                df.at[idx_to_edit, 'Importo'] = new_imp
                df.at[idx_to_edit, 'Periodo'] = new_per
                df.at[idx_to_edit, 'Descrizione'] = new_des
                df.to_csv(DATA_FILE, index=False)
                st.success("Modificata con successo!")
                st.rerun()
                
            if c_btn2.form_submit_button("🗑️ ELIMINA VOCE", use_container_width=True):
                df = df.drop(idx_to_edit)
                df.to_csv(DATA_FILE, index=False)
                st.warning("Voce eliminata.")
                st.rerun()

    # Visualizzazione Tabella semplice
    df_disp = df_f[['Data', 'Categoria', 'Descrizione', 'Importo', 'Periodo']].copy().sort_values("Data", ascending=False)
    df_disp['Data'] = df_disp['Data'].dt.strftime('%d/%m/%Y')
    st.dataframe(df_disp, use_container_width=True, hide_index=True)

else:
    st.info("Nessun dato trovato.")

# CONFIGURA MODELLI
st.divider()
with st.expander("⚙️ CONFIGURA MODELLI"):
    with st.form("f_mod"):
        c1 = st.selectbox("Categoria", cat_lista_clean)
        c2 = st.text_input("Nome")
        c3 = st.number_input("Importo", min_value=0.0)
        if st.form_submit_button("CREA"):
            nuovo_m = pd.DataFrame([[c1, c2, c3, "Mensile", date.today()]], columns=["Categoria", "Descrizione", "Importo", "Cadenza", "Data"])
            df_rec = pd.concat([df_rec, nuovo_m], ignore_index=True); df_rec.to_csv(RECURRING_FILE, index=False); st.rerun()
